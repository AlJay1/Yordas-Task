import pandas as pd
import openpyxl
import re


class Data_Process():
    def __init__(self):
        self.df = self.reading_the_dataframe()
        self.workbook = openpyxl.load_workbook('output.xlsx')
        self.worksheet = self.workbook['OUTPUT']

    def reading_the_dataframe(self):
        '''
        This function reads in the excel file as a data frame.
        '''
        df = pd.read_csv('GHS_v1.csv')
        # specify column names
        column_names = ['Index No.', 'Chemical name', 'CAS No.', 'Hazard class', 'Notes']
        df.columns = column_names
        # Remove the first row
        df = df.iloc[1:]

        df['Notes'] = df['Notes'].str.replace('\r', ' ')
        df['Chemical name'] = df['Chemical name'].str.replace('\r', ' ')

        # Replace all NaN values with 'N/A'
        df = df.fillna('N/A')
        df['Hazard class'] = df['Hazard class'].apply(lambda x: str(x))

        # extract number codes from each hazard class
        df['Number Codes'] = [re.findall(r'\d+[A-Z]?', hazard) for hazard in df['Hazard class']]
        return df
    
    def fill_excel(self):
        '''
        This function fills in the index, the cas number, the name and the notes of the chemical
        '''

        # Iterate through the rows of the DataFrame and write to the Excel file
        for index, row in self.df.iterrows():
            index_number = row["Index No."]
            cas_number = row['CAS No.']
            name = row['Chemical name']
            notes = row["Notes"]
            self.worksheet.cell(row=index+2, column=1).value = index_number
            self.worksheet.cell(row=index+2, column=2).value = cas_number
            self.worksheet.cell(row=index+2, column=3).value = name
            self.worksheet.cell(row=index+2, column=33).value = notes
        self.workbook.save('output.xlsx')
        

    def fill_hazards(self):
        '''
        This function fills in the hazards. It puts yes if they are present.
        '''        
        
        column_map = {
            'southafrica_ghs_2_1_display': 'Explos',
            'southafrica_ghs_2_2_display': 'Flam gas',
            'southafrica_ghs_2_3_display': 'Aeros',
            'southafrica_ghs_2_4_display': 'Oxid gas',
            'southafrica_ghs_2_5_display': 'Compr gas',
            'southafrica_ghs_2_6_display': 'Flam liq',
            'southafrica_ghs_2_7_display': 'Flam sol',
            'southafrica_ghs_2_8_display': 'Self reac',
            'southafrica_ghs_2_9_display': 'Pyro liq',
            'southafrica_ghs_2_10_display': 'Pyro sol',
            'southafrica_ghs_2_11_display': 'Self heat',
            'southafrica_ghs_2_12_display': 'Water-react',
            'southafrica_ghs_2_13_display': 'Oxid liq',
            'southafrica_ghs_2_14_display': 'Oxid sol',
            'southafrica_ghs_2_15_display': 'Org pero',
            'southafrica_ghs_2_16_display': 'Metal corr',
            'southafrica_ghs_2_17_display': 'Unstable expl',
            'southafrica_ghs_3_1_display': 'Acut tox',
            'southafrica_ghs_3_2_display': 'Skin corr',
            'southafrica_ghs_3_3_display': 'Eye corr',
            'southafrica_ghs_3_4_display': 'Skin sens',
            'southafrica_ghs_3_5_display': 'Mutag',
            'southafrica_ghs_3_6_display': 'Carci',
            'southafrica_ghs_3_7_display': 'Repro tox',
            'southafrica_ghs_3_8_display': 'STOT sing',
            'southafrica_ghs_3_9_display': 'STOT repe',
            'southafrica_ghs_3_10_display': 'Aspir haz',
            'southafrica_ghs_4_1_display': 'Chron aqua',
            'southafrica_ghs_4_2_display': 'Chron ozo',
        }

        # Loop through each row of the pandas dataframe
        for index, row in self.df.iterrows():
            # Retrieve the hazard information for the chemical
            hazard_info = row['Hazard class']
            # Update the corresponding columns in the worksheet
            for column, hazard in column_map.items():
                display = 'Yes' if hazard in hazard_info else ''
                self.worksheet.cell(row=index+2, column=int(self.worksheet.column_dimensions[column].index)).value = display

        self.workbook.save('output.xlsx')

    def replace_yes(self):
        '''
        This function replaces the yes' put in the fill hazards function with the number codes.
        '''
        df_v1 = pd.read_excel('output.xlsx')
        df_v1['Number Codes'] = self.df['Number Codes']
        codes = self.df['Number Codes'].tolist()
        for index, row in df_v1.iterrows():
            for i, col in enumerate(row):
                if col == 'Yes':
                    codes_row = codes[index]
                    if codes_row and len(codes_row) > 0:
                        code_value = f"Category {codes_row[0]}"
                        df_v1.iloc[index, i] = code_value
                    else:
                        df_v1.iloc[index, i] = ""

        df_v1.to_excel("mydataframe.xlsx", index=False)


if __name__ == "__main__":
    test = Data_Process()
    test.fill_excel()
    test.fill_hazards()
    test.replace_yes()